from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, date as date_type, timedelta, time
from .forms import *
from .models import *
from .planning_utils import generer_planning_journalier, build_timeline_data

# ============================================
# AUTHENTIFICATION
# ============================================

def home(request):
    """Page d'accueil - Login"""
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f"Bienvenue {user.first_name}!")
            return redirect('home')
        else:
            messages.error(request, "Identifiants invalides!")
            return redirect('home')
    else:
        return render(request, 'home.html')

def logout_user(request):
    """Déconnexion"""
    logout(request)
    messages.success(request, "Vous avez été déconnecté!")
    return redirect('home')

def register_user(request):
    """Inscription"""
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']
            user = authenticate(username=username, password=password)
            login(request, user)
            messages.success(request, "Inscription réussie! Bienvenue!")
            return redirect('dashboard')
    else:
        form = SignUpForm()
    return render(request, 'register.html', {'form': form})

# ============================================
# DASHBOARD ET PRODUCTION
# ============================================



#----------------------------Gamme opératoire--------------------------------------------------

@login_required(login_url='home')
def gammes_list(request):
    """
    Affiche la liste de toutes les gammes opératoires
    """
    gammes = GammeOperation.objects.all().order_by('ordre')
    context = {
        'gammes': gammes,
    }
    return render(request, 'gamme/gammes_list.html', context)


# Remplacer les 3 fonctions gammes par ces 2 seules fonctions

@login_required(login_url='home')
def gamme_form(request, id=None):
    """
    Ajoute ou modifie une gamme opératoire
    - Si id est None → Ajouter
    - Si id existe → Modifier
    """
    gamme = get_object_or_404(GammeOperation, id=id) if id else None
    titre = "Modifier la Gamme" if gamme else "Ajouter une Gamme"
    
    if request.method == 'POST':
        form = GammeOperationForm(request.POST, instance=gamme)
        if form.is_valid():
            form.save()
            msg = "modifiée" if gamme else "ajoutée"
            messages.success(request, f"✅ Gamme {msg} avec succès!")
            return redirect('gammes_list')
    else:
        form = GammeOperationForm(instance=gamme)
    
    return render(request, 'gamme/gamme_form.html', {'form': form, 'titre': titre, 'gamme': gamme})


@login_required(login_url='home')
def supprimer_gamme(request, id):
    """Supprime une gamme opératoire"""
    gamme = get_object_or_404(GammeOperation, id=id)
    gamme.delete()
    messages.success(request, "✅ Gamme supprimée!")
    return redirect('gammes_list')


@login_required(login_url='home')
def importer_gammes_excel(request):
    """
    Importe les gammes opératoires depuis un fichier Excel.
    Colonnes attendues: Nom | Temps alloué (heures) | Ordre
    """
    if request.method == 'POST':
        form = ImportGammesExcelForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = request.FILES['fichier_excel']
            
            try:
                # Importe openpyxl pour lire le fichier Excel
                from openpyxl import load_workbook
                
                # Charge le classeur Excel
                wb = load_workbook(fichier)
                ws = wb.active  # Feuille active
                
                gammes_creees = 0
                gammes_erreurs = []
                
                # Parcourt toutes les lignes (en commençant par la 2e, car la 1ère est l'en-tête)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Récupère les valeurs des 3 colonnes
                        nom = row[0]
                        temps_alloue = row[1]
                        ordre = row[2]
                        
                        # Valide que les champs ne sont pas vides
                        if not nom or temps_alloue is None or ordre is None:
                            gammes_erreurs.append(f"Ligne {row_idx}: données incomplètes")
                            continue
                        
                        # Convertit les valeurs au bon type
                        try:
                            temps_alloue = float(temps_alloue)
                            ordre = int(ordre)
                        except ValueError:
                            gammes_erreurs.append(
                                f"Ligne {row_idx}: temps ou ordre invalide (doit être numérique)"
                            )
                            continue
                        
                        # Crée la gamme opératoire
                        gamme, created = GammeOperation.objects.get_or_create(
                            ordre=ordre,
                            defaults={
                                'nom': str(nom),
                                'temps_alloue': temps_alloue,
                            }
                        )
                        
                        if created:
                            gammes_creees += 1
                        else:
                            # Si la gamme existe déjà, on la met à jour
                            gamme.nom = str(nom)
                            gamme.temps_alloue = temps_alloue
                            gamme.save()
                    
                    except Exception as e:
                        gammes_erreurs.append(f"Ligne {row_idx}: {str(e)}")
                
                # Affiche les messages
                if gammes_creees > 0:
                    messages.success(
                        request,
                        f"✅ {gammes_creees} gamme(s) importée(s) avec succès!"
                    )
                
                if gammes_erreurs:
                    for erreur in gammes_erreurs:
                        messages.warning(request, f"⚠️ {erreur}")
                
                return redirect('gammes_list')
            
            except Exception as e:
                messages.error(request, f"❌ Erreur lors de la lecture du fichier: {str(e)}")
                return redirect('importer_gammes_excel')
    
    else:
        form = ImportGammesExcelForm()
    
    return render(request, 'gamme/importer_gammes_excel.html', {'form': form})



#-----------------------------Opérateur--------------------------------------------------

@login_required(login_url='home')
def operateurs_list(request):
    """
    Affiche la liste de tous les opérateurs
    """
    operateurs = Operateur.objects.all().order_by('poste')

    context = {
        'operateurs': operateurs,
    }
    return render(request, 'operateur/operateur_list.html', context)


# Ajouter à la fin de views.py

@login_required(login_url='home')
def operateur_form(request, id=None):
    """
    Ajoute ou modifie un opérateur
    - Si id est None → Ajouter
    - Si id existe → Modifier
    """
    operateur = get_object_or_404(Operateur, pk=id) if id else None
    titre = "Modifier l'Opérateur" if operateur else "Ajouter un Opérateur"
    
    if request.method == 'POST':
        form = OperateurForm(request.POST, instance=operateur)
        if form.is_valid():
            form.save()
            msg = "modifié" if operateur else "ajouté"
            messages.success(request, f"✅ Opérateur {msg} avec succès!")
            return redirect('operateur_list')
    else:
        form = OperateurForm(instance=operateur)
    
    return render(request, 'operateur/operateur_form.html', {'form': form, 'titre': titre, 'operateur': operateur})


@login_required(login_url='home')
def supprimer_operateur(request, id):
    """Supprime un opérateur"""
    operateur = get_object_or_404(Operateur, pk=id)
    operateur.delete()
    messages.success(request, "✅ Opérateur supprimé!")
    return redirect('operateur_list')
    
#--------------------Ordre de fabrication ----------------------------------------------------------------------------

@login_required(login_url='home')
def of_list(request):
    """Affiche la liste de tous les OF"""
    ofs = OrdreFabrication.objects.all()
    context = {'ofs': ofs}
    return render(request, 'of/of_list.html', context)


@login_required(login_url='home')
def of_form(request, id=None):
    """Ajoute ou modifie un OF"""
    of = get_object_or_404(OrdreFabrication, id=id) if id else None
    titre = "Modifier l'OF" if of else "Ajouter un OF"
    
    if request.method == 'POST':
        form = OrdreFabricationForm(request.POST, instance=of)
        if form.is_valid():
            of = form.save()
            
            # 🔥 SI NOUVEL OF: créer automatiquement les 6 OperationOF
            if not id:  # C'est un nouvel OF
                # Récupérer toutes les opérations (gammes)
                gammes = GammeOperation.objects.all().order_by('ordre')
                for gamme in gammes:
                    # ✅ PLUS BESOIN d'assigner statut !
                    # statut est CALCULÉ AUTOMATIQUEMENT via @property
                    OperationOF.objects.create(
                        of=of,
                        gamme_operation=gamme
                        # ❌ SUPPRIMER: statut='en_attente'
                    )
                messages.success(request, f"✅ OF {of.numero} créé avec {gammes.count()} opérations!")
            else:
                messages.success(request, f"✅ OF {of.numero} modifié!")
            
            return redirect('of_list')
    else:
        form = OrdreFabricationForm(instance=of)
    
    return render(request, 'of/of_form.html', {'form': form, 'titre': titre, 'of': of})


@login_required(login_url='home')
def supprimer_of(request, id):
    """Supprime un OF"""
    of = get_object_or_404(OrdreFabrication, id=id)
    
    if request.method == 'POST':
        numero = of.numero
        of.delete()
        messages.success(request, f"✅ OF {numero} supprimé!")
        return redirect('of_list')
    
    return render(request, 'of/of_confirm_delete.html', {'of': of})


@login_required(login_url='home')
def of_detail(request, id):
    """Affiche le détail d'un OF avec sa timeline"""
    of = get_object_or_404(OrdreFabrication, id=id)
    operations = of.operationof_set.all().order_by('gamme_operation__ordre')
    
    context = {
        'of': of,
        'operations': operations,
        'avancement': of.avancement,
    }
    return render(request, 'of/of_detail.html', context)


@login_required(login_url='home')
def saisir_entree_operation(request, operation_id):
    """Saisit la date d'ENTRÉE d'une opération"""
    operation = get_object_or_404(OperationOF, id=operation_id)
    
    # Vérifier que l'opération n'a pas déjà commencé
    if operation.date_entree:
        messages.warning(request, "Cette opération a déjà une date d'entrée!")
        return redirect('of_detail', id=operation.of.id)
    
    if request.method == 'POST':
        form = EntreeOperationForm(request.POST, instance=operation)
        if form.is_valid():
            operation = form.save()  # ← Pas besoin de commit=False
            # ❌ PLUS BESOIN: operation.statut = 'en_cours'
            # ✅ Le statut est AUTOMATIQUEMENT 'en_cours' maintenant!
            
            messages.success(
                request,
                f"✅ {operation.gamme_operation.nom} — "
                f"Entrée: {operation.date_entree.strftime('%H:%M')} | "
                f"Sortie: {operation.date_sortie.strftime('%H:%M')} | "
                f"Statut: {operation.statut}"  # ← Affiche le statut CALCULÉ
            )
            return redirect('of_detail', id=operation.of.id)
    else:
        form = EntreeOperationForm(instance=operation)
    
    return render(request, 'of/saisir_entree.html', {
        'form': form,
        'operation': operation,
        'of': operation.of,
    })


@login_required(login_url='home')
def terminer_operation(request, operation_id):
    """
    Marque une opération comme TERMINÉE
    (OPTIONNEL MAINTENANT - elle se termine toute seule!)
    """
    operation = get_object_or_404(OperationOF, id=operation_id)
    
    if not operation.date_entree:
        messages.error(request, "Cette opération n'a pas de date d'entrée!")
        return redirect('of_detail', id=operation.of.id)
    
    # ❌ PLUS BESOIN: operation.statut = 'termine'
    # ✅ Le statut est AUTOMATIQUEMENT 'termine' si l'heure est atteinte!
    
    # On peut juste afficher un message de confirmation
    if operation.est_terminee:
        messages.success(
            request,
            f"✅ {operation.gamme_operation.nom} est TERMINÉE! "
            f"(Sortie: {operation.date_sortie.strftime('%H:%M')})"
        )
    else:
        remaining = operation.date_sortie - (timezone.now()+timedelta(hours=1))
        remaining_minutes = round(remaining.total_seconds() / 60)
        messages.info(
            request,
            f"⏳ {operation.gamme_operation.nom} — "
            f"Encore {remaining_minutes} minutes avant la fin..."
        )
    
    return redirect('of_detail', id=operation.of.id)



@login_required(login_url='home')
def list_retard(request):
    """
    Tableau simple des retards:
    - Lignes: Les 20 derniers OF
    - Colonnes: Les opérations (gammes)
    - Cellules: Le retard en minutes
    """
    # Les 20 derniers OF
    ofs = OrdreFabrication.objects.all().order_by('-id')[:20]
    
    # Toutes les gammes (opérations) triées par ordre
    gammes = GammeOperation.objects.all().order_by('ordre')
    
    context = {
        'ofs': ofs,
        'gammes': gammes,
    }
    return render(request, 'statistiques_retard/list_retard.html', context)


@login_required(login_url='home')
def dashboard(request):
    """
    Page d'accueil principal après connexion
    Affiche un résumé des OF
    """
    # ✅ SIMPLE - récupère juste tous les OF
    ofs = OrdreFabrication.objects.all()
    
    context = {
        'ofs': ofs,
        'total_of': ofs.count(),
        'today': datetime.now(),
    }
    return render(request, 'dashboard.html', context)




# ===== VUE: ASSIGNER LES GAMMES AVEC DRAG & DROP ================================================================================



@login_required(login_url='home')
def assigner_gammes_operateur(request, operateur_id):
    """
    Page Drag & Drop pour assigner les gammes opératoires à un opérateur.
    
    Interface simple:
    - À gauche: toutes les gammes disponibles (on peut les glisser)
    - À droite: zone de dépôt (les gammes assignées à cet opérateur)
    """
    # Récupère l'opérateur
    operateur = get_object_or_404(Operateur, id=operateur_id)
    
    if request.method == 'POST':
        # Traitement du formulaire POST (quand on soumet l'ordre des gammes)
        import json
        
        # Récupère les IDs des gammes dans l'ordre
        gammes_ids = request.POST.get('gammes_ordre', '[]')
        
        try:
            gammes_ids = json.loads(gammes_ids)
        except:
            gammes_ids = []
        
        # Supprime toutes les assignations existantes pour cet opérateur
        operateur.gammes.clear()
        
        # Ajoute les gammes dans le nouvel ordre
        for index, gamme_id in enumerate(gammes_ids):
            try:
                gamme = GammeOperation.objects.get(id=gamme_id)
                gamme.operateurs.add(operateur)
            except:
                pass
        
        messages.success(request, f"✅ Gammes assignées à {operateur}!")
        return redirect('dashboard_operateur', operateur_id=operateur_id)
    
    # GET: Affiche la page Drag & Drop
    # Récupère les gammes assignées à cet opérateur (dans l'ordre)
    gammes_assignees = operateur.gammes.all()
    
    # Récupère toutes les autres gammes (non assignées)
    gammes_disponibles = GammeOperation.objects.exclude(operateurs=operateur).order_by('ordre')
    
    context = {
        'operateur': operateur,
        'gammes_assignees': gammes_assignees,
        'gammes_disponibles': gammes_disponibles,
    }
    return render(request, 'planning/assigner_gammes_dragdrop.html', context)


# ============================================================
# VUE MODIFIER L'HEURE DE DÉBUT DU PLANNING
# ============================================================

@login_required(login_url='home')
def modifier_heure_debut(request, operateur_id):
    """
    Permet de modifier l'heure de début du planning pour un opérateur.
    Cela met à jour toutes les tâches du jour automatiquement.
    """
    operateur = get_object_or_404(Operateur, id=operateur_id)
    aujourd_hui = timezone.localdate() + timedelta(hours=1)
    
    # Récupère le planning du jour
    planning = PlanningJournalier.objects.filter(
        operateur=operateur,
        date=aujourd_hui
    ).first()
    
    if not planning:
        messages.error(request, "❌ Aucun planning pour aujourd'hui")
        return redirect('dashboard_operateur', operateur_id=operateur_id)
    
    if request.method == 'POST':
        # Récupère la nouvelle heure depuis le formulaire
        try:
            heure_str = request.POST.get('heure_debut', '')
            heure, minute = map(int, heure_str.split(':'))
            
            # Crée la nouvelle heure de début
            nouvelle_heure_debut = timezone.make_aware(
                datetime(aujourd_hui.year, aujourd_hui.month, aujourd_hui.day, heure, minute)
            )+ timedelta(hours=1)
            
            # Récupère toutes les tâches ordonnées
            taches = planning.taches.filter(type_tache='operation').order_by('ordre')
            
            # Recalcule les heures pour chaque tâche
            heure_courante = nouvelle_heure_debut
            
            for tache in taches:
                # Récupère la durée de la gamme
                duree = timedelta(hours=float(tache.gamme_operation.temps_alloue))
                
                # Met à jour les horaires
                tache.heure_debut_prevue = heure_courante
                tache.heure_fin_prevue = heure_courante + duree
                tache.save()
                
                # La prochaine tâche commence où finit celle-ci
                heure_courante = tache.heure_fin_prevue
            
            messages.success(request, f"✅ Heure de début modifiée à {heure:02d}:{minute:02d}")
            return redirect('dashboard_operateur', operateur_id=operateur_id)
        
        except Exception as e:
            messages.error(request, f"❌ Erreur: {str(e)}")
            return redirect('dashboard_operateur', operateur_id=operateur_id)
    
    # Affiche le formulaire avec l'heure actuelle
    heure_actuelle = planning.taches.filter(type_tache='operation').order_by('ordre').first()
    if heure_actuelle:
        heure_str = heure_actuelle.heure_debut_prevue.strftime('%H:%M')
    else:
        heure_str = '06:00'
    
    context = {
        'operateur': operateur,
        'planning': planning,
        'heure_str': heure_str,
    }
    return render(request, 'planning/modifier_heure_debut.html', context)


# ============================================================
# VUE SUPERVISEUR — Tous les opérateurs sur une seule page
# ============================================================

@login_required(login_url='home')
def superviseur_dashboard(request):
    if hasattr(request.user, 'operateur_profil') and request.user.operateur_profil is not None:
        messages.error(request, "🚫 Accès interdit aux opérateurs sur ce panneau.")
        return redirect('operateur_vue_simple')
    """
    Page superviseur : affiche tous les opérateurs avec leur timeline.
    Un opérateur par carte, avec sa timeline du shift.
    Rafraîchissement automatique toutes les 30 secondes (géré en JS).
    """
    aujourd_hui  = timezone.localdate() + timedelta(hours=1)
    operateurs   = Operateur.objects.all().order_by('poste','shift', 'nom')

    operateurs_data = []
    for operateur in operateurs:
        # Récupère le planning du jour pour cet opérateur (peut être None)
        planning = PlanningJournalier.objects.filter(
            operateur=operateur,
            date=aujourd_hui
        ).first()

        # Construit les données de timeline si le planning existe
        timeline_data = build_timeline_data(planning) if planning else None

        operateurs_data.append({
            'operateur'    : operateur,
            'planning'     : planning,
            'timeline_data': timeline_data,
            'has_gammes'   : GammeOperation.objects.filter(operateurs=operateur).exists(),
        })

    return render(request, 'planning/superviseur.html', {
        'operateurs_data': operateurs_data,
        'aujourd_hui'    : aujourd_hui,
    })


# ============================================================
# VUE DASHBOARD OPERATEUR — Timeline d'un seul opérateur
# ============================================================

@login_required(login_url='home')
def dashboard_operateur(request, operateur_id):
    """
    Page détaillée d'un opérateur avec sa timeline complète.
    Permet aussi de démarrer/terminer des tâches.
    """
    operateur   = get_object_or_404(Operateur,
     id=operateur_id)
    aujourd_hui = timezone.localdate() + timedelta(hours=1)

    # Planning du jour (peut être None si pas encore généré)
    planning = PlanningJournalier.objects.filter(
        operateur=operateur,
        date=aujourd_hui
    ).first()

    timeline_data = build_timeline_data(planning) if planning else None

    return render(request, 'planning/dashboard_operateur.html', {
        'operateur'    : operateur,
        'planning'     : planning,
        'timeline_data': timeline_data,
        'aujourd_hui'  : aujourd_hui,
        'has_gammes'   : GammeOperation.objects.filter(operateurs=operateur).exists(),
    })


@login_required(login_url='home')
def operateur_vue_simple(request):
    """
    Affiche toutes les tâches de l'opérateur pour la journée.
    L'opérateur peut cliquer sur n'importe quelle tâche pour la terminer.
    """
    # S'assurer que l'utilisateur est lié à un profil opérateur
    if not hasattr(request.user, 'operateur_profil'):
        messages.error(request, "Votre compte n'est pas associé à un profil opérateur.")
        return redirect('dashboard')

    # Récupère l'opérateur connecté
    operateur = request.user.operateur_profil
    today = timezone.now().date() + timedelta(hours=1)
    
    # Cherche le planning du jour pour cet opérateur
    planning = PlanningJournalier.objects.filter(operateur=operateur, date=today).first()
    
    # Liste de toutes les tâches (ordonnées par numéro d'ordre)
    taches = []
    if planning:
        taches = planning.taches.all().order_by('ordre')

    # Récupère les données du modal retard si présentes en session
    show_modal_retard = request.session.pop('show_modal_retard', False)
    tache_retard_id = request.session.pop('tache_retard_id', None)
    retard_minutes = request.session.pop('retard_minutes', None)
    tache_nom = request.session.pop('tache_nom', None)
    
    # Crée un formulaire pour le retard
    retard_form = RetardForm()
    tache_retard = None
    if tache_retard_id:
        tache_retard = get_object_or_404(TachePlanifiee, id=tache_retard_id)

    # Prépare le contexte pour la template
    context = {
        'operateur': operateur,
        'planning': planning,
        'taches': taches,
        'show_modal_retard': show_modal_retard,
        'retard_form': retard_form,
        'tache_retard': tache_retard,
        'retard_minutes': retard_minutes,
        'tache_nom': tache_nom,
    }
    
    return render(request, 'planning/operateur_vue_simple.html', context)

# ============================================================
# VUE GENERER PLANNING — Génère le planning du jour
# ============================================================

@login_required(login_url='home')
def generer_planning(request, operateur_id):
    """
    Génère (ou régénère) le planning du jour pour un opérateur.
    Redirige vers le dashboard opérateur après génération.
    """
    operateur   = get_object_or_404(Operateur, id=operateur_id)
    aujourd_hui = timezone.localdate() + timedelta(hours=1)

    # Appelle la fonction utilitaire qui crée les TachePlanifiee
    planning = generer_planning_journalier(operateur, aujourd_hui)

    messages.success(request, f"✅ Planning généré pour {operateur} ({planning.taches.count()} tâches)")
    return redirect('dashboard_operateur', operateur_id=operateur_id)



# ============================================================
# VUE TERMINER UNE TÂCHE
# ============================================================

@login_required(login_url='home')
def terminer_tache(request, tache_id):
    """
    Enregistre l'heure de fin RÉELLE d'une tâche.
    Si la tâche est en retard, affiche le modal de déclaration.
    Sinon, termine la tâche directement.
    """
    tache = get_object_or_404(TachePlanifiee, id=tache_id)

    if tache.heure_fin_reelle:
        messages.info(request, "Cette tâche est déjà marquée comme terminée.")
        return redirect('operateur_vue_simple')

    heure_fin_reelle = timezone.now()+ timedelta(hours=1)

    retard = heure_fin_reelle - tache.heure_fin_prevue

    # --- CAS 1 : La tâche est en retard ---
    if retard.total_seconds() > 1:
        retard_minutes = int(retard.total_seconds() / 60)
        
        # Stocke les données en session pour le modal
        request.session['tache_retard_id'] = tache_id
        request.session['retard_minutes'] = retard_minutes
        request.session['tache_nom'] = tache.nom_affichage
        request.session['show_modal_retard'] = True
        
        # Redirige vers la même page avec le modal affiché
        return redirect('operateur_vue_simple')

    # --- CAS 2 : La tâche est à l'heure ou en avance ---
    else:
        tache.heure_fin_reelle = heure_fin_reelle
        tache.save()
        messages.success(request, f"✅ Tâche '{tache.nom_affichage}' terminée dans les temps.")
        return redirect('operateur_vue_simple')




# ============================================================
# VUE POUR DÉCLARER UN RETARD (NOUVEAU)
# ============================================================
@login_required(login_url='home')
def declarer_retard(request, tache_id):
    """
    Affiche un formulaire pour que l'opérateur justifie un retard.
    Enregistre ensuite le retard et termine la tâche.
    """
    # Récupère la tâche concernée par le retard.
    tache = get_object_or_404(TachePlanifiee, id=tache_id)
    
    # Calcule la durée du retard en se basant sur l'heure actuelle.
    heure_actuelle = timezone.now() + timedelta(hours=1)
    retard_duree = heure_actuelle - tache.heure_fin_prevue

    # Convertit la durée du retard en minutes pour l'affichage.
    retard_minutes = int(retard_duree.total_seconds() / 60)

    # Si la méthode est POST, cela signifie que l'opérateur a soumis le formulaire.
    if request.method == 'POST':
        # Crée une instance du formulaire avec les données envoyées.
        form = RetardForm(request.POST)
        
        # Vérifie si le formulaire est valide (ex: une cause a bien été sélectionnée).
        if form.is_valid():
            # Marque la tâche comme terminée en enregistrant l'heure actuelle.
            tache.heure_fin_reelle = heure_actuelle
            tache.save()

            # Crée l'objet Retard mais ne le sauvegarde pas tout de suite en base de données.
            retard_obj = form.save(commit=False)
            retard_obj.tache = tache           # Associe le retard à la tâche.
            retard_obj.duree = retard_duree    # Enregistre la durée calculée du retard.
            retard_obj.save()                  # Sauvegarde l'objet Retard en base de données.

            # Affiche un message de succès.
            messages.success(request, "La cause du retard a bien été enregistrée.")
            
            # Redirige l'opérateur vers sa page de travail.
            return redirect('operateur_vue_simple')
    
    # Si la méthode est GET, on affiche simplement le formulaire vide.
    else:
        form = RetardForm()

    # Prépare le contexte à envoyer au template HTML.
    context = {
        'form': form,
        'tache': tache,
        'retard_minutes': retard_minutes,
    }
    # Affiche la page HTML du formulaire de retard.
    return render(request, 'planning/declarer_retard.html', context)



# ============================================================
# VUE RAFRÂICHIR OPÉRATEUR (supprimer toutes les opérations)
# ============================================================

@login_required(login_url='home')
def rafraichir_operateur(request, operateur_id):
    """
    Supprime toutes les gammes assignées à l'opérateur et son planning du jour.
    Utile pour recommencer à zéro.
    """
    operateur = get_object_or_404(Operateur, id=operateur_id)
    aujourd_hui = timezone.localdate() + timedelta(hours=1)

    # Récupère toutes les gammes assignées à cet opérateur
    gammes = operateur.gammes.all()
    gammes_supprimees = gammes.count()
    
    # Supprime l'assignation de toutes les gammes pour cet opérateur
    for gamme in gammes:
        gamme.operateurs.remove(operateur)

    # Supprime le planning du jour s'il existe
    planning_supprime = False
    planning = PlanningJournalier.objects.filter(operateur=operateur, date=aujourd_hui).first()
    if planning:
        planning.delete()
        planning_supprime = True

    # Message de confirmation
    message = f"✅ Opérateur rafraîchi : {gammes_supprimees} gammes supprimées"
    if planning_supprime:
        message += " et planning du jour supprimé"
    
    messages.success(request, message)
    return redirect('dashboard_operateur', operateur_id=operateur_id)


# ============================================================
# VUE DÉPLACER UNE TÂCHE (interchangeabilité)
# ============================================================

@login_required(login_url='home')
def deplacer_tache(request, tache_id, direction):
    """
    Permet de réordonner les tâches (monter ou descendre dans la liste).
    
    Direction: 'up' pour monter, 'down' pour descendre
    Cela change l'ordre de la tâche et recalcule les heures automatiquement.
    """
    # Récupère la tâche à déplacer
    tache = get_object_or_404(TachePlanifiee, id=tache_id)
    planning = tache.planning
    operateur = planning.operateur
    
    # Récupère toutes les tâches du planning (ordonnées)
    taches = list(planning.taches.filter(type_tache='operation').order_by('ordre'))
    
    # Trouve la position actuelle de la tâche
    index_actuel = None
    for i, t in enumerate(taches):
        if t.id == tache_id:
            index_actuel = i
            break
    
    if index_actuel is None:
        messages.error(request, "❌ Tâche non trouvée")
        return redirect('dashboard_operateur', operateur_id=operateur.id)
    
    # Échange avec la tâche voisine (up = vers le haut, down = vers le bas)
    if direction == 'up' and index_actuel > 0:
        # Échange avec la tâche précédente
        taches[index_actuel], taches[index_actuel - 1] = taches[index_actuel - 1], taches[index_actuel]
    elif direction == 'down' and index_actuel < len(taches) - 1:
        # Échange avec la tâche suivante
        taches[index_actuel], taches[index_actuel + 1] = taches[index_actuel + 1], taches[index_actuel]
    else:
        # Déjà au début (up) ou à la fin (down)
        messages.info(request, "ℹ️ Impossible de déplacer plus loin")
        return redirect('dashboard_operateur', operateur_id=operateur.id)
    
    # Recalcule les heures pour toutes les tâches
    h_debut, m_debut = operateur.shift_debut_hm
    heure_courante = timezone.make_aware(
        datetime(planning.date.year, planning.date.month, planning.date.day, h_debut, m_debut)
    ) + timedelta(hours=1)

    # Réassigne l'ordre et les heures
    for i, t in enumerate(taches):
        t.ordre = i + 1  # L'ordre recommence à 1
        t.heure_debut_prevue = heure_courante
        t.heure_fin_prevue = heure_courante + timedelta(hours=float(t.gamme_operation.temps_alloue))
        t.save()
        
        # La prochaine tâche commence où finit celle-ci
        heure_courante = t.heure_fin_prevue
    
    messages.success(request, f"✅ Tâches réordonnées")
    return redirect('dashboard_operateur', operateur_id=operateur.id)




# ============================================================
# DASHBOARD DIRECTEUR DE PRODUCTION
# ============================================================

@login_required(login_url='home')
def dashboard_directeur(request):
    """
    Dashboard pour le directeur de production.
    """
    
    # ============ 1. RÉCUPÉRER TOUS LES RETARDS ============
    tous_les_retards = Retard.objects.all()
    
    # ============ 2. PARETO PAR CAUSE ============
    retards_par_cause = {}
    
    for retard in tous_les_retards:
        nom_cause = retard.cause.nom
        
        if nom_cause not in retards_par_cause:
            retards_par_cause[nom_cause] = 0
        
        # Convertir en HEURES (diviser par 3600 secondes)
        retards_par_cause[nom_cause] += retard.duree.total_seconds() / 3600
    
    retards_par_cause_triee = dict(sorted(
        retards_par_cause.items(),
        key=lambda x: x[1],
        reverse=True
    ))
    
    causes_noms = list(retards_par_cause_triee.keys())
    causes_heures = list(retards_par_cause_triee.values())
    
    # ============ 3. PARETO PAR OPÉRATEUR ============
    retards_par_operateur = {}
    
    for retard in tous_les_retards:
        operateur = retard.tache.planning.operateur
        nom_operateur = f"{operateur.prenom} {operateur.nom}"
        
        if nom_operateur not in retards_par_operateur:
            retards_par_operateur[nom_operateur] = 0
        
        # Convertir en HEURES
        retards_par_operateur[nom_operateur] += retard.duree.total_seconds() / 3600
    
    retards_par_operateur_triee = dict(sorted(
        retards_par_operateur.items(),
        key=lambda x: x[1],
        reverse=True
    ))
    
    operateurs_noms = list(retards_par_operateur_triee.keys())
    operateurs_heures = list(retards_par_operateur_triee.values())
    
    # ============ 4. STATISTIQUES CUMULÉES ============
    
    # Total en HEURES
    total_retards_heures = sum(retard.duree.total_seconds() / 3600 for retard in tous_les_retards)
    
    nombre_total_retards = tous_les_retards.count()
    
    if nombre_total_retards > 0:
        retard_moyen_heures = total_retards_heures / nombre_total_retards
    else:
        retard_moyen_heures = 0
    
    retard_max_heures = 0
    if tous_les_retards.exists():
        retard_max_heures = max(retard.duree.total_seconds() / 3600 for retard in tous_les_retards)
    
    # ============ 5. STATISTIQUES PAR PÉRIODE ============
    
    aujourd_hui = timezone.now() + timedelta(hours=1)
    premier_jour_mois = aujourd_hui.replace(day=1)
    
    retards_mois = tous_les_retards
    retards_mois_heures = sum(retard.duree.total_seconds() / 3600 for retard in retards_mois)
    nombre_retards_mois = retards_mois.count()
    
    premier_jour_annee = aujourd_hui.replace(month=1, day=1)
    
    retards_annee = tous_les_retards
    retards_annee_heures = sum(retard.duree.total_seconds() / 3600 for retard in retards_annee)
    nombre_retards_annee = retards_annee.count()
    
    # ============ 6. PRÉPARER LE CONTEXTE ============
    context = {
        # Pareto par cause (en heures)
        'causes_noms': causes_noms,
        'causes_heures': causes_heures,
        
        # Pareto par opérateur (en heures)
        'operateurs_noms': operateurs_noms,
        'operateurs_heures': operateurs_heures,
        
        # Statistiques globales (en heures)
        'total_retards_heures': round(total_retards_heures, 2),
        'nombre_total_retards': nombre_total_retards,
        'retard_moyen_heures': round(retard_moyen_heures, 2),
        'retard_max_heures': round(retard_max_heures, 2),
        
        # Statistiques mois (en heures)
        'retards_mois_heures': round(retards_mois_heures, 2),
        'nombre_retards_mois': nombre_retards_mois,
        
        # Statistiques année (en heures)
        'retards_annee_heures': round(retards_annee_heures, 2),
        'nombre_retards_annee': nombre_retards_annee,
        
        # Date
        'aujourd_hui': aujourd_hui,
    }
    
    return render(request, 'directeur/dashboard_directeur.html', context)






